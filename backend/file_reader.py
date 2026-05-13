"""File ingestion for Amé — PDF / Excel / Word / image / text.

Single public entry point: `read_file(path_or_token, prompt=None)`. The
dispatcher routes by detected file kind (extension + magic bytes, which
must agree — mismatch means quarantine). All extracted text is wrapped
through `untrusted.wrap` before it's returned to the model.

Tokens issued by the `/upload` endpoint are opaque UUIDs registered via
`register_token(token, real_path, filename)`. Tools only ever see the
token, never the absolute path — keeps prompt-injected "read /etc/shadow"
attempts out of reach of the model.
"""

from __future__ import annotations
import re
import threading
import uuid
from dataclasses import dataclass
from pathlib import Path

from backend import untrusted
from backend import path_guard


_token_lock = threading.Lock()
_tokens: dict[str, dict] = {}


def register_token(path: str, filename: str | None = None, kind: str = "") -> str:
    """Register a real path under an opaque token. Returns the token."""
    token = uuid.uuid4().hex[:16]
    with _token_lock:
        _tokens[token] = {"path": str(Path(path).expanduser()), "filename": filename or Path(path).name, "kind": kind}
    return token


def resolve_token(token: str) -> dict | None:
    with _token_lock:
        return _tokens.get(token)


def forget_token(token: str) -> None:
    with _token_lock:
        _tokens.pop(token, None)


# ── Caps ───────────────────────────────────────────────────────────────────
MAX_PDF_PAGES        = 20
MAX_XLSX_PREVIEW_ROWS = 50
MAX_XLSX_FULL_ROWS    = 500
MAX_TEXT_CHARS        = 200_000
MAX_FILE_BYTES        = 26_214_400  # 25 MB

_TEXT_EXTS = {
    ".c", ".h", ".cs", ".go", ".js", ".md", ".py", ".rb", ".rs", ".sh", ".ts",
    ".bat", ".cpp", ".css", ".csv", ".hpp", ".ini", ".jsx", ".log", ".php",
    ".ps1", ".tsx", ".txt", ".xml", ".yml", ".html", ".java", ".json", ".scss",
    ".toml", ".yaml",
}
_IMAGE_EXTS = {".bmp", ".gif", ".jpg", ".png", ".jpeg", ".webp"}
_MAGIC = {
    "pdf":  (b"%PDF-",),
    "xlsx": (b"PK\x03\x04",),
    "docx": (b"PK\x03\x04",),
    "png":  (b"\x89PNG\r\n\x1a\n",),
    "jpg":  (b"\xff\xd8\xff",),
    "gif":  (b"GIF87a", b"GIF89a"),
    "webp": (b"RIFF",),
    "bmp":  (b"BM",),
}


def _sniff_kind(path: Path, head: bytes) -> str:
    """Return a normalized kind: pdf / xlsx / docx / image / text / mismatch / unknown."""
    ext = path.suffix.lower()
    # Image extensions check magic strictly.
    if ext in _IMAGE_EXTS:
        for k in (ext[1:], "jpg" if ext == ".jpeg" else ext[1:]):
            for sig in _MAGIC.get(k, ()):
                if head.startswith(sig):
                    return "image"
        return "mismatch"
    if ext == ".pdf":
        return "pdf" if any(head.startswith(s) for s in _MAGIC["pdf"]) else "mismatch"
    if ext == ".xlsx":
        return "xlsx" if any(head.startswith(s) for s in _MAGIC["xlsx"]) else "mismatch"
    if ext == ".docx":
        return "docx" if any(head.startswith(s) for s in _MAGIC["docx"]) else "mismatch"
    if ext in _TEXT_EXTS:
        return "text"
    return "unknown"


@dataclass
class Quarantine:
    reason: str


def _quarantine_check(path: Path, kind: str, raw: bytes) -> Quarantine | None:
    """Return a Quarantine reason, or None if the file is clean enough to read."""
    try:
        size = path.stat().st_size
    except Exception:
        return Quarantine(reason="Cannot stat file.")
    if size > MAX_FILE_BYTES:
        return Quarantine(reason=f"File is too large ({size} bytes, cap {MAX_FILE_BYTES}).")
    if kind == "mismatch":
        return Quarantine(reason="File extension doesn't match its contents — refusing to read.")
    if kind == "pdf" and re.search(rb"/JavaScript\b|/JS\b|/OpenAction\b", raw[:4_194_304]):
        return Quarantine(reason="PDF contains embedded JavaScript or auto-open actions.")
    return None


def read_pdf(path: Path) -> dict:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as e:
        return {"success": False, "error": f"pypdf not installed: {e}"}
    try:
        reader = PdfReader(str(path))
    except Exception as e:
        return {"success": False, "error": f"PDF open failed: {e}"}
    chunks: list[str] = []
    for i, page in enumerate(reader.pages[:MAX_PDF_PAGES]):
        try:
            chunks.append(page.extract_text() or "")
        except Exception:
            chunks.append("")
    text = "\n\n".join(c for c in chunks if c.strip())
    return {"success": True, "text": text[:MAX_TEXT_CHARS], "pages": len(reader.pages)}


def read_excel(path: Path, sheet: str | None = None, full: bool = False) -> dict:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        return {"success": False, "error": f"openpyxl not installed: {e}"}
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        return {"success": False, "error": f"xlsx open failed: {e}"}
    sheets = wb.sheetnames
    target_sheet = sheet if (sheet and sheet in sheets) else sheets[0]
    ws = wb[target_sheet]
    limit = MAX_XLSX_FULL_ROWS if full else MAX_XLSX_PREVIEW_ROWS
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= limit:
            break
        rows.append(list(row))
    counts = {s: ws.max_row for s in sheets[:8]}
    return {"success": True, "sheet": target_sheet, "sheets": sheets,
            "rows": rows, "row_counts": counts}


def read_docx(path: Path) -> dict:
    try:
        from docx import Document  # type: ignore
    except Exception as e:
        return {"success": False, "error": f"python-docx not installed: {e}"}
    try:
        doc = Document(str(path))
    except Exception as e:
        return {"success": False, "error": f"docx open failed: {e}"}
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs)
    return {"success": True, "text": text[:MAX_TEXT_CHARS]}


def read_image(path: Path, prompt: str | None = None) -> dict:
    try:
        with open(path, "rb") as f:
            data = f.read()
    except Exception as e:
        return {"success": False, "error": f"image read failed: {e}"}
    return {"success": True, "kind": "image", "bytes": data,
            "mime_type": _guess_mime(path), "path": str(path),
            "filename": path.name, "prompt": prompt or ""}


def read_text(path: Path) -> dict:
    try:
        with path.open("r", encoding="utf-8", errors="replace") as f:
            text = f.read(MAX_TEXT_CHARS + 1)
    except Exception as e:
        return {"success": False, "error": f"text read failed: {e}"}
    truncated = len(text) > MAX_TEXT_CHARS
    return {"success": True, "text": text[:MAX_TEXT_CHARS], "truncated": truncated}


def _guess_mime(path: Path) -> str:
    ext = path.suffix.lower()
    return {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp"}.get(ext, "application/octet-stream")


def _read_head(path: Path) -> bytes:
    try:
        with path.open("rb") as f:
            return f.read(4096)
    except Exception:
        return b""


def read_file(path_or_token: str, prompt: str | None = None) -> dict:
    """Resolve a token or path and route to the correct reader."""
    if not path_or_token:
        return {"success": False, "error": "No path or token provided."}

    info = resolve_token(path_or_token)
    if info:
        real_path = Path(info["path"])
        friendly_source = f"upload:{info.get('filename', 'file')}"
    else:
        real_path = Path(path_or_token).expanduser().resolve()
        friendly_source = f"file:{real_path.name}"
        guard = path_guard.guard_or_error(real_path, mode="read")
        if guard is not None:
            return guard

    if not real_path.exists() or not real_path.is_file():
        return {"success": False, "error": f"File not found: {real_path}"}

    head = _read_head(real_path)
    kind = _sniff_kind(real_path, head)
    quar = _quarantine_check(real_path, kind, head)
    if quar:
        return {"success": False, "error": quar.reason, "quarantined": True}

    if kind == "image":
        return read_image(real_path, prompt=prompt)

    if kind == "pdf":
        result = read_pdf(real_path)
    elif kind == "xlsx":
        result = read_excel(real_path, full=False)
    elif kind == "docx":
        result = read_docx(real_path)
    elif kind == "text":
        result = read_text(real_path)
    else:
        return {"success": False, "error": f"Unsupported file kind: {kind}"}

    if result.get("success") and "text" in result:
        result["text"] = untrusted.wrap(friendly_source, result["text"], max_chars=MAX_TEXT_CHARS)
    return result


AUTO_READ_PREVIEW_CHARS = 8000


def resolve_image_bytes(path_or_token: str) -> dict:
    """Resolve an upload token → raw image bytes + MIME."""
    if not path_or_token:
        return {"success": False, "error": "No path or token provided."}
    info = resolve_token(path_or_token)
    if info:
        real_path = Path(info["path"])
    else:
        real_path = Path(path_or_token).expanduser().resolve()
        guard = path_guard.guard_or_error(real_path, mode="read")
        if guard is not None:
            return guard
    if not real_path.exists() or not real_path.is_file():
        return {"success": False, "error": f"File not found: {real_path}"}
    head = _read_head(real_path)
    kind = _sniff_kind(real_path, head)
    if kind != "image":
        return {"success": False, "error": f"Not an image (kind={kind})"}
    quar = _quarantine_check(real_path, kind, head)
    if quar:
        return {"success": False, "error": quar.reason, "quarantined": True}
    try:
        data = real_path.read_bytes()
    except Exception as e:
        return {"success": False, "error": f"read failed: {e}"}
    return {"success": True, "bytes": data, "mime_type": _guess_mime(real_path),
            "kind": "image", "filename": real_path.name, "path": str(real_path)}


def auto_read_preview(path_or_token: str) -> dict:
    """Read a file and return a short preview for auto-injection after upload."""
    result = read_file(path_or_token)
    if not result.get("success"):
        return result
    text = result.get("text", "") or ""
    if len(text) > AUTO_READ_PREVIEW_CHARS:
        result["text"] = text[:AUTO_READ_PREVIEW_CHARS] + "\n[...preview truncated — ask for 'read the full file' to see more...]"
        result["preview"] = True
    else:
        result["preview"] = False
    return result
